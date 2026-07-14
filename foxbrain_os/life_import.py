"""Reviewed SAP export imports for the FireFox Living Enterprise."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import time
import uuid
from pathlib import Path

from .living_enterprise import (
    ensure_living_enterprise_schema,
    record_timeline_event,
    relate_life_objects,
    upsert_life_object,
)


MAX_IMPORT_BYTES = 50 * 1024 * 1024
IMPORT_ROLES = {"boss", "admin"}
SHARED_ACCOUNT_NAMES = {
    "店长", "收银员", "总经理", "店助", "采购部", "财务部", "总部", "管理员", "admin",
}
NORMALIZED_SHARED_ACCOUNT_NAMES = {item.lower() for item in SHARED_ACCOUNT_NAMES}

OBJECT_CONFIGS = {
    "supplier_life": {
        "label": "供应商生命",
        "code": "supplier_code",
        "name": "supplier_name",
        "mapping": {
            "supplier_code": ["业务伙伴代码", "供应商编号", "供应商代码", "CardCode"],
            "supplier_name": ["业务伙伴名称", "供应商名称", "CardName"],
            "contact": ["联系人", "联系人姓名"],
            "phone": ["电话 1", "电话1", "电话", "手机"],
            "phone_secondary": ["电话 2", "电话2"],
            "address": ["地址", "供应商地址"],
            "group_name": ["组名称", "供应商组"],
            "account_balance": ["科目余额", "余额"],
            "cooperation_start": ["合作时间", "合作开始时间"],
        },
        "recommended": ["contact", "phone", "address", "cooperation_start"],
    },
    "product_life": {
        "label": "产品生命",
        "code": "product_code",
        "name": "product_name",
        "mapping": {
            "product_code": ["物料编号", "商品编码", "SKU", "ItemCode"],
            "product_name": ["物料描述", "商品名称", "产品名称", "ItemName"],
            "brand": ["制造商名称", "品牌", "品牌名称"],
            "category": ["组名称", "类别", "商品类别", "品类"],
            "inventory": ["存货量", "库存", "库存数量", "OnHand"],
            "purchase_price": ["采购价", "采购价格"],
            "member_price": ["会员价", "零售价", "销售价"],
            "barcode": ["附加标识", "条码", "国际条码"],
            "sales_history": ["销售历史"],
            "sales_trend": ["销售趋势"],
            "use_scenario": ["使用场景"],
        },
        "recommended": ["brand", "category"],
    },
    "people_life": {
        "label": "人才生命",
        "code": "employee_code",
        "name": "employee_name",
        "mapping": {
            "employee_code": ["员工编号", "员工代码", "FUserCode", "EmpID"],
            "employee_name": ["姓名", "员工姓名", "FUserName", "FirstName"],
            "store": ["门店", "所属门店", "分店"],
            "position": ["岗位", "职位", "职务"],
            "status": ["状态", "在职状态"],
            "hire_date": ["入职时间", "入职日期"],
            "training_records": ["培训记录"],
            "capability_tags": ["能力标签"],
            "growth_path": ["成长路径"],
        },
        "recommended": ["store", "position", "status", "hire_date"],
    },
}


def _now():
    return int(time.time())


def _json(value):
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return "" if text.upper() in {"NULL", "NONE", "N/A"} else text


def _number(value):
    text = _clean(value).replace(",", "")
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1]
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return None


def _safe_filename(filename):
    name = Path(str(filename or "import.xls")).name
    cleaned = re.sub(r"[^0-9A-Za-z._\-\u4e00-\u9fff]+", "_", name).strip("._")
    return cleaned[:180] or "import.xls"


def ensure_life_import_schema(conn):
    ensure_living_enterprise_schema(conn)
    conn.execute(
        """create table if not exists life_import_batches(
        id integer primary key autoincrement,
        batch_id text not null unique,
        object_type text not null,
        original_filename text not null,
        stored_path text not null,
        file_hash text not null,
        file_size integer not null,
        file_format text not null,
        source_time text,
        sheet_name text,
        status text not null,
        row_count integer not null default 0,
        valid_count integer not null default 0,
        warning_count integer not null default 0,
        error_count integer not null default 0,
        imported_count integer not null default 0,
        updated_count integer not null default 0,
        skipped_count integer not null default 0,
        mapping_json text not null default '{}',
        validation_summary_json text not null default '{}',
        created_by integer,
        approved_by integer,
        created_at integer not null,
        reviewed_at integer,
        imported_at integer,
        rolled_back_at integer,
        updated_at integer not null)"""
    )
    conn.execute(
        """create table if not exists life_import_rows(
        id integer primary key autoincrement,
        batch_id text not null,
        row_number integer not null,
        row_hash text not null,
        raw_json text not null,
        mapped_json text not null,
        validation_json text not null default '[]',
        status text not null,
        life_id text,
        created_at integer not null,
        updated_at integer not null,
        unique(batch_id,row_number))"""
    )
    conn.execute(
        """create table if not exists life_import_changes(
        id integer primary key autoincrement,
        batch_id text not null,
        life_id text not null,
        action text not null,
        previous_json text,
        imported_json text not null,
        created_at integer not null,
        unique(batch_id,life_id))"""
    )
    conn.execute(
        """create table if not exists life_import_logs(
        id integer primary key autoincrement,
        batch_id text not null,
        event_type text not null,
        actor_id integer,
        detail_json text not null default '{}',
        created_at integer not null)"""
    )
    for statement in (
        "create index if not exists idx_life_import_batches_status on life_import_batches(status,created_at)",
        "create index if not exists idx_life_import_batches_hash on life_import_batches(file_hash,object_type)",
        "create index if not exists idx_life_import_rows_status on life_import_rows(batch_id,status,row_number)",
        "create index if not exists idx_life_import_changes_batch on life_import_changes(batch_id,life_id)",
        "create index if not exists idx_life_import_logs_batch on life_import_logs(batch_id,created_at)",
    ):
        conn.execute(statement)


def _parse_delimited(data, filename):
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        text = data.decode("utf-16")
        encoding = "utf-16"
    else:
        text = None
        encoding = ""
        for candidate in ("utf-8-sig", "gb18030"):
            try:
                text = data.decode(candidate)
                encoding = candidate
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("文件编码无法识别，请导出为 UTF-8、UTF-16 或 GBK。")
    delimiter = "\t" if "\t" in text.partition("\n")[0] else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matrix = [[_clean(value) for value in row] for row in reader]
    return matrix, "{}-delimited".format(encoding), ""


def _parse_xls(data):
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("服务器缺少旧版 Excel 解析组件 xlrd。") from exc
    workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    sheet = next((workbook.sheet_by_index(i) for i in range(workbook.nsheets)
                  if workbook.sheet_by_index(i).nrows), None)
    if sheet is None:
        raise ValueError("Excel 文件没有可导入的数据。")
    matrix = [[_clean(sheet.cell_value(row, col)) for col in range(sheet.ncols)]
              for row in range(sheet.nrows)]
    return matrix, "xls-biff", sheet.name


def _parse_xlsx(data):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("服务器缺少新版 Excel 解析组件 openpyxl。") from exc
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = next((item for item in workbook.worksheets if item.max_row), None)
    if sheet is None:
        raise ValueError("Excel 文件没有可导入的数据。")
    matrix = [[_clean(value) for value in row] for row in sheet.iter_rows(values_only=True)]
    return matrix, "xlsx", sheet.title


def parse_export_bytes(data, filename):
    if not data:
        raise ValueError("上传文件为空。")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("文件超过 50MB，未进入导入流程。")
    suffix = Path(filename).suffix.lower()
    if data.startswith(b"\xd0\xcf\x11\xe0"):
        matrix, file_format, sheet_name = _parse_xls(data)
    elif data.startswith(b"PK\x03\x04"):
        matrix, file_format, sheet_name = _parse_xlsx(data)
    elif suffix in {".xls", ".csv", ".tsv", ".txt"} or data.startswith((b"\xff\xfe", b"\xfe\xff")):
        matrix, file_format, sheet_name = _parse_delimited(data, filename)
    else:
        raise ValueError("仅支持 SAP 导出的 XLS、XLSX、CSV 或制表符文本。")
    matrix = [row for row in matrix if any(_clean(value) for value in row)]
    if len(matrix) < 2:
        raise ValueError("文件没有可导入的数据行。")
    headers = [_clean(value) or "未命名字段{}".format(index + 1) for index, value in enumerate(matrix[0])]
    rows = []
    for values in matrix[1:]:
        padded = values + [""] * max(0, len(headers) - len(values))
        rows.append({headers[index]: padded[index] for index in range(len(headers))})
    return {"headers": headers, "rows": rows, "file_format": file_format, "sheet_name": sheet_name}


def resolve_mapping(headers, object_type):
    config = OBJECT_CONFIGS.get(object_type)
    if not config:
        raise ValueError("不支持的生命对象类型。")
    normalized = {re.sub(r"\s+", "", header).lower(): header for header in headers}
    result = {}
    for target, aliases in config["mapping"].items():
        source = ""
        for alias in aliases:
            source = normalized.get(re.sub(r"\s+", "", alias).lower(), "")
            if source:
                break
        result[target] = source
    return result


def map_and_validate_rows(conn, object_type, headers, rows, mapping_override=None):
    config = OBJECT_CONFIGS[object_type]
    mapping = resolve_mapping(headers, object_type)
    if mapping_override is not None:
        mapping = {
            target: _clean(mapping_override.get(target, ""))
            for target in config["mapping"]
        }
        unknown = sorted({source for source in mapping.values() if source and source not in headers})
        if unknown:
            raise ValueError("字段映射包含源文件中不存在的列：{}".format("、".join(unknown)))
    seen_codes = set()
    output = []
    for index, raw in enumerate(rows, start=2):
        mapped = {target: _clean(raw.get(source, "")) if source else "" for target, source in mapping.items()}
        issues = []
        code = mapped[config["code"]]
        name = mapped[config["name"]]
        if not code:
            issues.append({"level": "error", "code": "missing_code", "message": "编号为空"})
        if not name:
            issues.append({"level": "error", "code": "missing_name", "message": "名称为空"})
        if code and code in seen_codes:
            issues.append({"level": "error", "code": "duplicate_code", "message": "文件内编号重复"})
        seen_codes.add(code)
        for field in config["recommended"]:
            if not mapped.get(field):
                issues.append({"level": "warning", "code": "missing_" + field, "message": "缺少{}".format(field)})
        if object_type == "people_life" and name.strip().lower() in NORMALIZED_SHARED_ACCOUNT_NAMES:
            issues.append({"level": "warning", "code": "possible_shared_account", "message": "可能是岗位共用账号，需人工确认"})
        numeric_fields = ("account_balance",) if object_type == "supplier_life" else (
            ("inventory", "purchase_price", "member_price") if object_type == "product_life" else ()
        )
        for field in numeric_fields:
            if mapped.get(field) and _number(mapped[field]) is None:
                issues.append({"level": "warning", "code": "invalid_" + field, "message": "数值格式无法识别"})
        existing = None
        if code:
            existing = conn.execute(
                "select life_id from living_objects where object_type=? and object_ref_type='sap_export' and object_ref_id=?",
                (object_type, code),
            ).fetchone()
        if existing:
            issues.append({"level": "warning", "code": "existing_object", "message": "已有生命对象，确认后生成新版本"})
        level = "error" if any(item["level"] == "error" for item in issues) else (
            "warning" if issues else "valid"
        )
        output.append({
            "row_number": index,
            "raw": raw,
            "mapped": mapped,
            "issues": issues,
            "status": level,
            "row_hash": hashlib.sha256(_json(raw).encode("utf-8")).hexdigest(),
        })
    return mapping, output


def update_import_mapping(conn, batch_id, mapping, actor_id=None):
    """Apply a reviewed header mapping and re-run validation without touching Life Objects."""
    ensure_life_import_schema(conn)
    batch = conn.execute("select * from life_import_batches where batch_id=?", (batch_id,)).fetchone()
    if not batch:
        raise ValueError("导入批次不存在。")
    if batch["status"] != "review_required":
        raise ValueError("只有等待人工确认的批次可以调整字段映射。")
    summary = json.loads(batch["validation_summary_json"] or "{}")
    headers = summary.get("headers") or []
    raw_rows = []
    stored_rows = conn.execute(
        "select id,row_number,row_hash,raw_json from life_import_rows where batch_id=? order by row_number",
        (batch_id,),
    ).fetchall()
    for row in stored_rows:
        raw_rows.append(json.loads(row["raw_json"] or "{}"))
    resolved, checked = map_and_validate_rows(conn, batch["object_type"], headers, raw_rows, mapping)
    now = _now()
    conn.executemany(
        """update life_import_rows set mapped_json=?,validation_json=?,status=?,life_id=null,updated_at=?
        where id=?""",
        [(_json(item["mapped"]), _json(item["issues"]), item["status"], now, stored_rows[index]["id"])
         for index, item in enumerate(checked)],
    )
    valid_count = sum(1 for row in checked if row["status"] == "valid")
    warning_count = sum(1 for row in checked if row["status"] == "warning")
    error_count = sum(1 for row in checked if row["status"] == "error")
    summary.update({"valid": valid_count, "warnings": warning_count, "errors": error_count})
    conn.execute(
        """update life_import_batches set valid_count=?,warning_count=?,error_count=?,mapping_json=?,
        validation_summary_json=?,updated_at=? where batch_id=?""",
        (valid_count, warning_count, error_count, _json(resolved), _json(summary), now, batch_id),
    )
    result = {"valid": valid_count, "warnings": warning_count, "errors": error_count, "sap_write": False}
    _log(conn, batch_id, "mapping_reviewed", actor_id, result)
    return result


def _log(conn, batch_id, event_type, actor_id, detail=None):
    conn.execute(
        "insert into life_import_logs(batch_id,event_type,actor_id,detail_json,created_at) values(?,?,?,?,?)",
        (batch_id, event_type, actor_id, _json(detail or {}), _now()),
    )


def stage_import_bytes(conn, data, filename, object_type, storage_root, actor_id=None, source_time=""):
    ensure_life_import_schema(conn)
    if object_type not in OBJECT_CONFIGS:
        raise ValueError("请选择正确的生命对象类型。")
    digest = hashlib.sha256(data).hexdigest()
    duplicate = conn.execute(
        """select * from life_import_batches where file_hash=? and object_type=?
        and status not in ('rolled_back','failed') order by id desc limit 1""",
        (digest, object_type),
    ).fetchone()
    if duplicate:
        return {"ok": True, "duplicate": True, "batch": dict(duplicate)}
    parsed = parse_export_bytes(data, filename)
    mapping, rows = map_and_validate_rows(conn, object_type, parsed["headers"], parsed["rows"])
    batch_id = "LIFE-IMPORT-" + uuid.uuid4().hex[:16].upper()
    root = Path(storage_root) / "originals" / time.strftime("%Y%m%d") / batch_id
    root.mkdir(parents=True, exist_ok=False)
    stored_path = root / _safe_filename(filename)
    stored_path.write_bytes(data)
    valid_count = sum(1 for row in rows if row["status"] == "valid")
    warning_count = sum(1 for row in rows if row["status"] == "warning")
    error_count = sum(1 for row in rows if row["status"] == "error")
    now = _now()
    summary = {
        "headers": parsed["headers"],
        "valid": valid_count,
        "warnings": warning_count,
        "errors": error_count,
        "confirmation_required": True,
        "sap_write": False,
    }
    conn.execute(
        """insert into life_import_batches(
        batch_id,object_type,original_filename,stored_path,file_hash,file_size,file_format,source_time,sheet_name,
        status,row_count,valid_count,warning_count,error_count,mapping_json,validation_summary_json,
        created_by,created_at,updated_at) values(?,?,?,?,?,?,?,?,?,'review_required',?,?,?,?,?,?,?,?,?)""",
        (batch_id, object_type, Path(filename).name, str(stored_path), digest, len(data), parsed["file_format"],
         source_time, parsed["sheet_name"], len(rows), valid_count, warning_count, error_count,
         _json(mapping), _json(summary), actor_id, now, now),
    )
    conn.executemany(
        """insert into life_import_rows(
        batch_id,row_number,row_hash,raw_json,mapped_json,validation_json,status,created_at,updated_at)
        values(?,?,?,?,?,?,?,?,?)""",
        [(batch_id, row["row_number"], row["row_hash"], _json(row["raw"]), _json(row["mapped"]),
          _json(row["issues"]), row["status"], now, now) for row in rows],
    )
    _log(conn, batch_id, "file_staged_for_review", actor_id, summary)
    batch = conn.execute("select * from life_import_batches where batch_id=?", (batch_id,)).fetchone()
    return {"ok": True, "duplicate": False, "batch": dict(batch)}


def _life_payload_for_row(object_type, mapped):
    if object_type == "supplier_life":
        return {
            "code": mapped["supplier_code"], "name": mapped["supplier_name"],
            "identity": {"supplier_code": mapped["supplier_code"], "supplier_name": mapped["supplier_name"]},
            "origin": {"source": "SAP B1 export", "cooperation_start": mapped.get("cooperation_start", "")},
            "state": {"contact": mapped.get("contact", ""), "phone": mapped.get("phone", ""),
                      "phone_secondary": mapped.get("phone_secondary", ""), "address": mapped.get("address", ""),
                      "group_name": mapped.get("group_name", ""),
                      "account_balance": _number(mapped.get("account_balance"))},
            "future": {"status": "awaiting_confirmed_business_plan"},
        }
    if object_type == "product_life":
        return {
            "code": mapped["product_code"], "name": mapped["product_name"],
            "identity": {"product_code": mapped["product_code"], "product_name": mapped["product_name"],
                         "barcode": mapped.get("barcode", "")},
            "origin": {"source": "SAP B1 export", "brand": mapped.get("brand", ""),
                       "category": mapped.get("category", "")},
            "state": {"inventory": _number(mapped.get("inventory")),
                      "purchase_price": _number(mapped.get("purchase_price")),
                      "member_price": _number(mapped.get("member_price")),
                      "sales_history": mapped.get("sales_history", ""), "sales_trend": mapped.get("sales_trend", "")},
            "future": {"use_scenario": mapped.get("use_scenario", ""), "status": "awaiting_evidence"},
        }
    return {
        "code": mapped["employee_code"], "name": mapped["employee_name"],
        "identity": {"employee_code": mapped["employee_code"], "employee_name": mapped["employee_name"],
                     "identity_center_link": "pending"},
        "origin": {"source": "SAP B1 export", "hire_date": mapped.get("hire_date", "")},
        "state": {"store": mapped.get("store", ""), "position": mapped.get("position", ""),
                  "status": mapped.get("status", ""), "training_records": mapped.get("training_records", ""),
                  "capability_tags": mapped.get("capability_tags", "")},
        "future": {"growth_path": mapped.get("growth_path", ""), "status": "requires_identity_review"},
    }


def _row_source(batch, row):
    return {
        "source_type": "sap_export",
        "source_id": batch["batch_id"],
        "source_ref": "{}#row={}".format(batch["original_filename"], row["row_number"]),
        "source_time": batch["source_time"] or batch["created_at"],
        "evidence": [{"row": row["row_number"], "hash": row["row_hash"], "file_hash": batch["file_hash"]}],
    }


def _link_confirmed_relationships(conn, batch, row, life_id, object_type, mapped):
    source = _row_source(batch, row)
    if object_type == "product_life":
        candidates = []
        if mapped.get("brand"):
            candidates.append(("brand_life", mapped["brand"], "belongs_to_brand"))
        for target_type, name, relation in candidates:
            target = conn.execute(
                "select life_id from living_objects where object_type=? and trim(display_name)=trim(?) and status='active'",
                (target_type, name),
            ).fetchone()
            if target:
                relate_life_objects(conn, life_id, target[0], relation, **source)
    elif object_type == "people_life" and mapped.get("store"):
        target = conn.execute(
            "select life_id from living_objects where object_type='store_life' and trim(display_name)=trim(?) and status='active'",
            (mapped["store"],),
        ).fetchone()
        if target:
            relate_life_objects(conn, life_id, target[0], "works_at", **source)


def approve_import(conn, batch_id, actor_id=None, include_warnings=False):
    ensure_life_import_schema(conn)
    batch_row = conn.execute("select * from life_import_batches where batch_id=?", (batch_id,)).fetchone()
    if not batch_row:
        raise ValueError("导入批次不存在。")
    batch = dict(batch_row)
    if batch["status"] != "review_required":
        raise ValueError("该批次当前不能确认导入。")
    if batch["error_count"]:
        raise ValueError("批次仍有错误行，请修正源文件后重新上传。")
    if batch["warning_count"] and not include_warnings:
        raise ValueError("批次存在警告，必须勾选确认后才能导入。")
    rows = conn.execute(
        "select * from life_import_rows where batch_id=? and status in ('valid','warning') order by row_number",
        (batch_id,),
    ).fetchall()
    created = updated = skipped = 0
    now = _now()
    for row_value in rows:
        row = dict(row_value)
        mapped = json.loads(row["mapped_json"])
        payload = _life_payload_for_row(batch["object_type"], mapped)
        existing_row = conn.execute(
            """select * from living_objects where object_type=? and object_ref_type='sap_export'
            and object_ref_id=?""",
            (batch["object_type"], payload["code"]),
        ).fetchone()
        previous = dict(existing_row) if existing_row else None
        source = _row_source(batch, row)
        result = upsert_life_object(
            conn, batch["object_type"], "sap_export", payload["code"], payload["name"],
            identity=payload["identity"], origin=payload["origin"], state=payload["state"], future=payload["future"],
            **source
        )
        record_timeline_event(
            conn, result["life_id"], "sap_export_confirmed", "人工确认 SAP 导出资料",
            batch["source_time"] or batch["created_at"], "来源文件：{}".format(batch["original_filename"]), **source
        )
        _link_confirmed_relationships(conn, batch, row, result["life_id"], batch["object_type"], mapped)
        current = dict(conn.execute("select * from living_objects where life_id=?", (result["life_id"],)).fetchone())
        conn.execute(
            """insert into life_import_changes(batch_id,life_id,action,previous_json,imported_json,created_at)
            values(?,?,?,?,?,?)""",
            (batch_id, result["life_id"], "created" if previous is None else "updated",
             _json(previous) if previous else None, _json(current), now),
        )
        conn.execute(
            "update life_import_rows set status='imported',life_id=?,updated_at=? where id=?",
            (result["life_id"], now, row["id"]),
        )
        created += 1 if previous is None else 0
        updated += 1 if previous is not None and result["changed"] else 0
        skipped += 1 if previous is not None and not result["changed"] else 0
    conn.execute(
        """update life_import_batches set status='imported',approved_by=?,reviewed_at=?,imported_at=?,
        imported_count=?,updated_count=?,skipped_count=?,updated_at=? where batch_id=?""",
        (actor_id, now, now, created, updated, skipped, now, batch_id),
    )
    result = {"created": created, "updated": updated, "skipped": skipped, "sap_write": False}
    _log(conn, batch_id, "import_approved", actor_id, result)
    return result


def rollback_import(conn, batch_id, actor_id=None):
    ensure_life_import_schema(conn)
    batch = conn.execute("select * from life_import_batches where batch_id=?", (batch_id,)).fetchone()
    if not batch or batch["status"] != "imported":
        raise ValueError("只有已导入且未回滚的批次可以回滚。")
    changes = conn.execute("select * from life_import_changes where batch_id=? order by id desc", (batch_id,)).fetchall()
    for item in changes:
        current = conn.execute("select * from living_objects where life_id=?", (item["life_id"],)).fetchone()
        imported = json.loads(item["imported_json"])
        if not current or int(current["version"]) != int(imported["version"]):
            raise ValueError("生命对象在导入后已有新版本，已阻止自动回滚。")
        if item["action"] == "created":
            other_source = conn.execute(
                """select 1 from living_object_sources where life_id=?
                and not (source_type='sap_export' and source_id=?) limit 1""",
                (item["life_id"], batch_id),
            ).fetchone()
            if other_source:
                raise ValueError("新生命对象已有其他来源，已阻止删除。")
    conn.execute("delete from living_relationships where source_type='sap_export' and source_id=?", (batch_id,))
    conn.execute("delete from living_timeline_events where source_type='sap_export' and source_id=?", (batch_id,))
    conn.execute("delete from living_object_sources where source_type='sap_export' and source_id=?", (batch_id,))
    for item in changes:
        if item["action"] == "created":
            conn.execute("delete from living_objects where life_id=?", (item["life_id"],))
        else:
            previous = json.loads(item["previous_json"])
            conn.execute(
                """update living_objects set display_name=?,identity_json=?,origin_json=?,state_json=?,future_json=?,
                status=?,version=?,created_at=?,updated_at=? where life_id=?""",
                (previous["display_name"], previous["identity_json"], previous["origin_json"],
                 previous["state_json"], previous["future_json"], previous["status"], previous["version"],
                 previous["created_at"], previous["updated_at"], item["life_id"]),
            )
    now = _now()
    conn.execute("update life_import_rows set status='rolled_back',updated_at=? where batch_id=?", (now, batch_id))
    conn.execute(
        "update life_import_batches set status='rolled_back',rolled_back_at=?,updated_at=? where batch_id=?",
        (now, now, batch_id),
    )
    _log(conn, batch_id, "import_rolled_back", actor_id, {"objects": len(changes), "sap_write": False})
    return {"rolled_back": len(changes), "sap_write": False}


def import_batch_detail(conn, batch_id, row_limit=100):
    ensure_life_import_schema(conn)
    batch = conn.execute("select * from life_import_batches where batch_id=?", (batch_id,)).fetchone()
    if not batch:
        return {"ok": False, "message": "导入批次不存在"}
    data = dict(batch)
    for field in ("mapping_json", "validation_summary_json"):
        data[field[:-5]] = json.loads(data.pop(field) or "{}")
    rows = []
    for row in conn.execute(
        "select * from life_import_rows where batch_id=? order by row_number limit ?", (batch_id, int(row_limit))
    ).fetchall():
        item = dict(row)
        for field in ("raw_json", "mapped_json", "validation_json"):
            item[field[:-5]] = json.loads(item.pop(field) or ("[]" if field == "validation_json" else "{}"))
        rows.append(item)
    data["rows"] = rows
    data["logs"] = [dict(row) for row in conn.execute(
        "select * from life_import_logs where batch_id=? order by created_at desc", (batch_id,)
    ).fetchall()]
    return {"ok": True, "batch": data}


def list_import_batches(conn, limit=50):
    ensure_life_import_schema(conn)
    return [dict(row) for row in conn.execute(
        "select * from life_import_batches order by created_at desc,id desc limit ?", (int(limit),)
    ).fetchall()]


def life_object_allowed(user_role, object_type, object_state=None, user_store=""):
    role = str(user_role or "")
    if role in {"boss", "admin"}:
        return True
    if object_type == "supplier_life":
        return role in {"purchasing", "finance"}
    if object_type == "product_life":
        return role in {"purchasing", "store_manager", "employee", "finance"}
    if object_type == "people_life":
        state = object_state or {}
        return role == "store_manager" and bool(user_store) and state.get("store") == user_store
    if object_type == "store_life":
        return role in {"store_manager", "purchasing", "finance"}
    if object_type == "brand_life":
        return role in {"store_manager", "purchasing", "finance", "employee"}
    return False
