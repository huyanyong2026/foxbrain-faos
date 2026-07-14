import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from apps.ai.operation import (
    CORE_TABLES,
    OPERATION_SCHEMA_STATEMENTS,
    CoreSnapshotClient,
    export_inventory_health,
    export_replenishment,
    inventory_health_analysis,
    replenishment_analysis,
)


ROOT = Path(__file__).resolve().parents[1]
AI_ROOT = ROOT / "apps" / "ai"


def sample_snapshot():
    return {
        "OWHS": [{"WhsCode": "NS", "WhsName": "南山店"}],
        "OITB": [{"ItmsGrpCod": 1, "ItmsGrpNam": "背包"}],
        "OMRC": [{"FirmCode": 10, "FirmName": "KAILAS"}],
        "OITM": [
            {"ItemCode": "A", "ItemName": "轻量背包", "ItmsGrpCod": 1, "FirmCode": 10},
            {"ItemCode": "B", "ItemName": "经典背包", "ItmsGrpCod": 1, "FirmCode": 10},
            {"ItemCode": "C", "ItemName": "长期库存", "ItmsGrpCod": 1, "FirmCode": 10},
        ],
        "OITW": [
            {"ItemCode": "A", "WhsCode": "NS", "OnHand": 3, "IsCommited": 0, "OnOrder": 2, "AvgPrice": 100},
            {"ItemCode": "B", "WhsCode": "NS", "OnHand": 4, "IsCommited": 0, "OnOrder": 0, "AvgPrice": 200},
            {"ItemCode": "C", "WhsCode": "NS", "OnHand": 2, "IsCommited": 0, "OnOrder": 0, "AvgPrice": 300},
        ],
        "OINV": [
            {"DocEntry": 1, "DocDate": "2026-07-01", "CANCELED": "N"},
            {"DocEntry": 2, "DocDate": "2026-06-01", "CANCELED": "N"},
            {"DocEntry": 3, "DocDate": "2025-12-25", "CANCELED": "N"},
        ],
        "INV1": [
            {"DocEntry": 1, "ItemCode": "A", "WhsCode": "NS", "Quantity": 10},
            {"DocEntry": 2, "ItemCode": "A", "WhsCode": "NS", "Quantity": 4},
            {"DocEntry": 3, "ItemCode": "B", "WhsCode": "NS", "Quantity": 1},
        ],
        "ORIN": [],
        "RIN1": [],
    }


class FakeConnector:
    def __init__(self, snapshot):
        self.snapshot = snapshot
        self.paths = []

    def get_json(self, path):
        self.paths.append(path)
        if path == "api/v1/status":
            return {"ok": True, "data": {"mode": "read_only", "mirror": {"status": "completed"}}}
        if path.startswith("api/v1/operation/snapshot"):
            return {"ok": True, "data": {
                "warehouse": {"WhsCode": "NS", "WhsName": "南山店"},
                "products": self.snapshot.get("products", []), "sales": self.snapshot.get("sales", []),
            }}
        table = path.split("/")[4]
        query = path.split("?", 1)[1]
        values = dict(part.split("=") for part in query.split("&"))
        offset = int(values["offset"])
        limit = int(values["limit"])
        return {"ok": True, "data": {"rows": self.snapshot.get(table, [])[offset:offset + limit]}}


class AiOperationCenterTest(unittest.TestCase):
    def test_schema_preserves_runs_and_audit(self):
        schema = "\n".join(OPERATION_SCHEMA_STATEMENTS)
        self.assertIn("operation_analysis_runs", schema)
        self.assertIn("operation_analysis_audit", schema)
        self.assertIn("source_updated_at", schema)
        self.assertIn("pending_review", schema)

    def test_replenishment_formula_and_evidence(self):
        result = replenishment_analysis(sample_snapshot(), "南山店", date(2026, 7, 13))
        item = next(row for row in result["items"] if row["product_code"] == "A")
        self.assertEqual(item["sales_30"], 10)
        self.assertEqual(item["sales_60"], 14)
        self.assertEqual(item["suggested_quantity"], 5)
        self.assertEqual(item["sales_trend"], "上升")
        self.assertTrue(result["human_confirmation_required"])
        self.assertEqual(set(result["source_tables"]), set(CORE_TABLES))

    def test_inventory_health_thresholds(self):
        result = inventory_health_analysis(sample_snapshot(), "NS", date(2026, 7, 13))
        by_code = {row["product_code"]: row for row in result["items"]}
        self.assertEqual(by_code["A"]["health_level"], "健康")
        self.assertEqual(by_code["B"]["health_level"], "橙色")
        self.assertEqual(by_code["C"]["health_level"], "黑色")
        self.assertEqual(result["inventory_amount"], 1700)
        self.assertTrue(result["human_confirmation_required"])

    def test_excel_exports_are_readable(self):
        replenishment = replenishment_analysis(sample_snapshot(), "南山店", date(2026, 7, 13))
        inventory = inventory_health_analysis(sample_snapshot(), "南山店", date(2026, 7, 13))
        with tempfile.TemporaryDirectory() as directory:
            replenishment_path = Path(directory) / "补货.xlsx"
            inventory_path = Path(directory) / "库存.xlsx"
            export_replenishment(replenishment, replenishment_path)
            export_inventory_health(inventory, inventory_path)
            self.assertEqual(load_workbook(replenishment_path).active["A1"].value, "门店")
            self.assertEqual(load_workbook(inventory_path).active["L1"].value, "AI建议")

    def test_core_client_uses_only_get_endpoints_and_paginates(self):
        connector = FakeConnector({"OITM": [{"ItemCode": "A"}, {"ItemCode": "B"}, {"ItemCode": "C"}]})
        client = CoreSnapshotClient(connector, page_size=2)
        self.assertEqual(len(client.table("OITM")), 3)
        self.assertTrue(all(path.startswith("api/v1/") for path in connector.paths))

    def test_compact_core_snapshot_avoids_full_sales_download(self):
        connector = FakeConnector({
            "products": [{"ItemCode": "A", "ItemName": "背包", "OnHand": 3, "OnOrder": 2}],
            "sales": [{"ItemCode": "A", "Sales30": 10, "Sales60": 14, "Sales90": 14, "Sales180": 14}],
        })
        snapshot = CoreSnapshotClient(connector).operation_snapshot("南山店", date(2026, 7, 13))
        result = replenishment_analysis(snapshot, "南山店", date(2026, 7, 13))
        self.assertEqual(result["items"][0]["suggested_quantity"], 5)
        self.assertEqual(len(connector.paths), 1)

    def test_routes_are_scoped_and_no_sap_connection_exists(self):
        source = (AI_ROOT / "app.py").read_text(encoding="utf-8")
        operation = (AI_ROOT / "operation.py").read_text(encoding="utf-8")
        for marker in (
            "/operation/replenishment", "/operation/inventory-health",
            "operation_store(current_user()", "pending_review", "CoreSnapshotClient",
        ):
            self.assertIn(marker, source)
        combined = source + operation
        self.assertNotIn("47.107.117.131", combined)
        self.assertNotIn("pytds.connect", combined)
        self.assertNotIn("UPDATE OITW", combined.upper())


if __name__ == "__main__":
    unittest.main()
