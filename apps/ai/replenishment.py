"""Deterministic replenishment rules and import/export helpers.

This module never connects to SAP.  It accepts normalized facts supplied by
Enterprise Data Core or by an administrator-uploaded SAP export.
"""

from __future__ import annotations

import csv
import io
import math
import os
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo


RULE_VERSION = "replenishment-v1.0"
TRIGGER_DAYS = int(os.environ.get("REPLENISHMENT_TRIGGER_DAYS", "15"))
TARGET_DAYS = int(os.environ.get("REPLENISHMENT_TARGET_DAYS", "30"))
URGENT_DAYS = int(os.environ.get("REPLENISHMENT_URGENT_DAYS", "7"))
BUSINESS_TIMEZONE = os.environ.get("REPLENISHMENT_TIMEZONE", "Asia/Shanghai")
BUSINESS_TZ = ZoneInfo(BUSINESS_TIMEZONE)
ALLOWED_STORES = {
    "nanshan": "南山店",
    "hangyuan": "航苑店",
    "zhenxing": "振兴店",
}


def business_now():
    return datetime.now(BUSINESS_TZ)


def business_today():
    return business_now().date()
STORE_ALIASES = {
    "南山": "nanshan", "南山店": "nanshan", "nanshan": "nanshan",
    "航苑": "hangyuan", "航苑店": "hangyuan", "hangyuan": "hangyuan",
    "振兴": "zhenxing", "振兴店": "zhenxing", "zhenxing": "zhenxing",
}


FIELD_ALIASES = {
    "store_code": ("store_code", "门店编码", "仓库编码"),
    "store_name": ("store_name", "门店", "门店名称", "仓库名称"),
    "sku_code": ("sku_code", "商品编码", "货号", "itemcode"),
    "product_name": ("product_name", "商品名称", "品名", "itemname"),
    "brand_name": ("brand_name", "品牌", "品牌名称"),
    "category_name": ("category_name", "产品类别", "品类", "类别"),
    "color": ("color", "颜色"),
    "size": ("size", "尺码", "规格"),
    "available_stock": ("available_stock", "available_qty", "当前库存", "可售库存"),
    "sales_30d": ("sales_30d", "近30天销售", "最近30天销售"),
    "sales_prev_30d": ("sales_prev_30d", "前30天销售", "31-60天销售"),
    "sales_60d": ("sales_60d", "近60天销售", "最近60天销售"),
}


@dataclass(frozen=True)
class ReplenishmentResult:
    store_code: str
    store_name: str
    sku_code: str
    product_name: str
    brand_name: str
    category_name: str
    color: str
    size: str
    available_stock: int
    sales_30d: int
    sales_prev_30d: int
    sales_60d: int
    avg_daily_sales: float
    stock_days: float | None
    safety_stock: int
    suggested_qty: int
    priority: str
    recommendation: str
    reason: str
    warning: str


def _number(value, field, row_number):
    if value in (None, ""):
        return Decimal(0)
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        raise ValueError("第 {} 行的{}不是有效数字".format(row_number, field))


def _integer(value, field, row_number):
    number = _number(value, field, row_number)
    if number != number.to_integral_value():
        raise ValueError("第 {} 行的{}必须是整数件数".format(row_number, field))
    return int(number)


def _value(row, canonical):
    normalized = {str(key).strip().lower(): value for key, value in row.items() if key is not None}
    for alias in FIELD_ALIASES[canonical]:
        if alias.lower() in normalized:
            return normalized[alias.lower()]
    return None


def normalize_store(store_code, store_name=""):
    for value in (store_code, store_name):
        key = str(value or "").strip().lower()
        if key in STORE_ALIASES:
            code = STORE_ALIASES[key]
            return code, ALLOWED_STORES[code]
    raise ValueError("门店必须是南山店、航苑店或振兴店")


def normalize_input_rows(raw_rows):
    """Normalize a flat Core/file dataset and reject unusable records."""
    normalized = []
    seen = set()
    errors = []
    for row_number, raw in enumerate(raw_rows, start=2):
        try:
            store_code, store_name = normalize_store(_value(raw, "store_code"), _value(raw, "store_name"))
            sku_code = str(_value(raw, "sku_code") or "").strip()
            product_name = str(_value(raw, "product_name") or "").strip()
            if not sku_code or not product_name:
                raise ValueError("商品编码和商品名称不能为空")
            if _value(raw, "available_stock") is None or _value(raw, "sales_30d") is None:
                raise ValueError("当前库存和近30天销售不能为空")
            key = (store_code, sku_code)
            if key in seen:
                raise ValueError("同一门店商品编码重复：{}".format(sku_code))
            seen.add(key)
            sales_30d = _integer(_value(raw, "sales_30d"), "近30天销售", row_number)
            previous_value = _value(raw, "sales_prev_30d")
            sales_60_value = _value(raw, "sales_60d")
            if previous_value in (None, "") and sales_60_value in (None, ""):
                raise ValueError("必须提供前30天销售或近60天销售")
            if previous_value in (None, "") and sales_60_value not in (None, ""):
                sales_prev_30d = max(0, _integer(sales_60_value, "近60天销售", row_number) - sales_30d)
            else:
                sales_prev_30d = _integer(previous_value, "前30天销售", row_number)
            normalized.append({
                "store_code": store_code,
                "store_name": store_name,
                "sku_code": sku_code,
                "product_name": product_name,
                "brand_name": str(_value(raw, "brand_name") or "").strip(),
                "category_name": str(_value(raw, "category_name") or "").strip(),
                "color": str(_value(raw, "color") or "").strip(),
                "size": str(_value(raw, "size") or "").strip(),
                "available_stock": _integer(_value(raw, "available_stock"), "当前库存", row_number),
                "sales_30d": sales_30d,
                "sales_prev_30d": sales_prev_30d,
            })
        except ValueError as exc:
            errors.append(str(exc) if str(exc).startswith("第 ") else "第 {} 行：{}".format(row_number, exc))
    if errors:
        preview = "；".join(errors[:10])
        suffix = "；另有 {} 项错误".format(len(errors) - 10) if len(errors) > 10 else ""
        raise ValueError(preview + suffix)
    if not normalized:
        raise ValueError("文件中没有可用的补货数据")
    return normalized


def calculate_replenishment(item):
    stock_raw = int(item.get("available_stock") or 0)
    stock = max(0, stock_raw)
    sales_30d_raw = int(item.get("sales_30d") or 0)
    sales_previous_raw = int(item.get("sales_prev_30d") or 0)
    sales_30d = max(0, sales_30d_raw)
    sales_previous = max(0, sales_previous_raw)
    sales_60d = sales_30d + sales_previous
    warning_parts = []
    if stock_raw < 0:
        warning_parts.append("源数据为负库存，计算时按0件处理")
    if sales_30d_raw < 0 or sales_previous_raw < 0:
        warning_parts.append("源数据存在负净销量，计算时按0件处理")

    avg_daily = sales_30d / 30.0
    stock_days = stock / avg_daily if avg_daily > 0 else None
    safety_stock = math.ceil(avg_daily * TRIGGER_DAYS)
    target_stock = math.ceil(avg_daily * TARGET_DAYS)
    suggested_qty = max(0, target_stock - stock)

    if sales_60d == 0:
        priority = "不补"
        recommendation = "do_not_replenish"
        suggested_qty = 0
        reason = "该商品近60天无销售，当前库存{}件，暂不建议补货。".format(stock)
    elif stock_days is not None and stock_days < TRIGGER_DAYS:
        priority = "紧急" if stock_days < URGENT_DAYS else "高"
        if priority == "高" and sales_30d > sales_previous:
            priority = "紧急"
        recommendation = "replenish"
        growth = "，近30天销量高于前30天" if sales_30d > sales_previous else ""
        reason = (
            "该商品近30天销售{}件，当前可售库存{}件，按近期销量预计约{}天售罄{}，建议补货{}件。"
            .format(sales_30d, stock, max(0, round(stock_days)), growth, suggested_qty)
        )
    else:
        priority = "普通" if suggested_qty > 0 else "不补"
        recommendation = "review" if suggested_qty > 0 else "do_not_replenish"
        days_text = "暂无近期销量" if stock_days is None else "预计可售约{}天".format(round(stock_days))
        reason = "该商品近30天销售{}件，当前库存{}件，{}，暂不列入紧急补货。".format(
            sales_30d, stock, days_text
        )

    return ReplenishmentResult(
        store_code=item["store_code"], store_name=item["store_name"], sku_code=item["sku_code"],
        product_name=item["product_name"], brand_name=item.get("brand_name", ""),
        category_name=item.get("category_name", ""), color=item.get("color", ""), size=item.get("size", ""),
        available_stock=stock_raw, sales_30d=sales_30d, sales_prev_30d=sales_previous,
        sales_60d=sales_60d, avg_daily_sales=round(avg_daily, 4),
        stock_days=None if stock_days is None else round(stock_days, 2), safety_stock=safety_stock,
        suggested_qty=suggested_qty, priority=priority, recommendation=recommendation,
        reason=reason, warning="；".join(warning_parts),
    )


def calculate_batch(rows):
    priority_order = {"紧急": 0, "高": 1, "普通": 2, "不补": 3}
    results = [asdict(calculate_replenishment(row)) for row in rows]
    return sorted(results, key=lambda row: (row["store_name"], priority_order[row["priority"]], row["stock_days"] if row["stock_days"] is not None else 10**9, row["sku_code"]))


def parse_uploaded_file(filename, content):
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        return normalize_input_rows(csv.DictReader(io.StringIO(text)))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("服务器缺少 Excel 解析组件") from exc
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, ())]
        return normalize_input_rows(dict(zip(headers, values)) for values in rows if any(value not in (None, "") for value in values))
    raise ValueError("只支持 .xlsx 或 .csv 文件")


EXPORT_HEADERS = (
    ("门店", "store_name"), ("商品编码", "sku_code"), ("商品名称", "product_name"),
    ("品牌", "brand_name"), ("产品类别", "category_name"), ("颜色", "color"), ("尺码", "size"),
    ("当前库存", "available_stock"), ("近30天销售", "sales_30d"), ("前30天销售", "sales_prev_30d"),
    ("近60天销售", "sales_60d"), ("日均销售", "avg_daily_sales"), ("库存可销售天数", "stock_days"),
    ("安全库存", "safety_stock"), ("建议补货数量", "suggested_qty"), ("优先级", "priority"),
    ("AI建议原因", "reason"), ("数据警告", "warning"),
)


def build_excel(items, metadata):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    book = Workbook()
    sheet = book.active
    sheet.title = "补货建议"
    sheet.append([label for label, _key in EXPORT_HEADERS])
    for item in items:
        sheet.append([item.get(key) if item.get(key) is not None else "" for _label, key in EXPORT_HEADERS])
    header_fill = PatternFill("solid", fgColor="176443")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [12, 16, 30, 16, 16, 12, 12, 12, 14, 14, 14, 12, 16, 12, 16, 12, 52, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        row[16].alignment = Alignment(wrap_text=True, vertical="top")
        row[17].alignment = Alignment(wrap_text=True, vertical="top")

    info = book.create_sheet("生成说明")
    info.append(["项目", "内容"])
    for key, value in metadata.items():
        info.append([key, value])
    info.append(["规则说明", "库存可销售天数低于15天触发；目标30天；低于7天为紧急；60天无销售不补货"])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 80
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    stream = io.BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream


def new_batch_id(source, business_date=None):
    value = business_date or business_today().isoformat()
    timestamp = business_now().strftime("%H%M%S")
    return "{}-{}-{}-{}".format(source, value.replace("-", ""), timestamp, uuid.uuid4().hex[:6])
