"""Evidence-backed replenishment and inventory-health analysis."""

from __future__ import annotations

import math
import os
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


OPERATION_SCHEMA_STATEMENTS = (
    """create table if not exists operation_analysis_runs(
    id bigserial primary key,run_id varchar(80) unique not null,analysis_type varchar(40) not null,
    store_id bigint not null,store_name varchar(160) not null,source_type varchar(80) not null default 'core.vafox.com',
    source_ref text not null,source_updated_at timestamptz,result_json jsonb not null default '[]'::jsonb,
    summary_json jsonb not null default '{}'::jsonb,excel_path text,status varchar(30) not null default 'pending_review',
    created_by bigint not null,approved_by bigint,approved_at timestamptz,created_at timestamptz not null default now())""",
    """create table if not exists operation_analysis_audit(
    id bigserial primary key,run_id varchar(80) not null,action varchar(40) not null,user_id bigint,
    note text not null default '',created_at timestamptz not null default now())""",
    "create index if not exists idx_operation_runs_type_store on operation_analysis_runs(analysis_type,store_id,created_at desc)",
    "create index if not exists idx_operation_runs_creator on operation_analysis_runs(created_by,created_at desc)",
)

CORE_TABLES = ("OWHS", "OITM", "OITB", "OMRC", "OITW", "OINV", "INV1", "ORIN", "RIN1")


def _field(row, *names, default=None):
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
        lower = name.lower()
        for key, value in row.items():
            if str(key).lower() == lower and value not in (None, ""):
                return value
    return default


def _number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _day(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value)[:10]
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def _round_quantity(value):
    return max(0, int(math.ceil(float(value or 0))))


class CoreSnapshotClient:
    def __init__(self, connector, page_size=200, max_rows=300000):
        self.connector = connector
        self.page_size = min(200, max(1, int(page_size)))
        self.max_rows = int(max_rows)

    def table(self, name):
        rows = []
        offset = 0
        while offset < self.max_rows:
            result = self.connector.get_json(
                "api/v1/tables/dbo/{}/rows?limit={}&offset={}".format(name, self.page_size, offset)
            )
            if not result.get("ok"):
                raise RuntimeError("Core API读取{}失败：{}".format(name, result.get("error", "未知错误")))
            batch = (result.get("data") or {}).get("rows") or []
            rows.extend(batch)
            if len(batch) < self.page_size:
                break
            offset += len(batch)
        if offset >= self.max_rows:
            raise RuntimeError("Core API数据量超过单次分析安全上限")
        return rows

    def snapshot(self, tables=CORE_TABLES):
        return {name: self.table(name) for name in tables}

    def source_status(self):
        result = self.connector.get_json("api/v1/status")
        if not result.get("ok"):
            raise RuntimeError("Core API状态读取失败：{}".format(result.get("error", "未知错误")))
        return result.get("data") or {}

    def operation_snapshot(self, store_reference, as_of=None):
        as_of = as_of or date.today()
        result = self.connector.get_json(
            "api/v1/operation/snapshot?store={}&as_of={}".format(quote(str(store_reference)), as_of.isoformat())
        )
        if not result.get("ok"):
            raise RuntimeError("Core API门店经营快照读取失败：{}".format(result.get("error", "未知错误")))
        data = result.get("data") or {}
        return {
            "COMPACT_WAREHOUSE": data.get("warehouse") or {},
            "COMPACT_PRODUCTS": data.get("products") or [],
            "COMPACT_SALES": data.get("sales") or [],
        }


def resolve_warehouse(snapshot, store_reference):
    if snapshot.get("COMPACT_WAREHOUSE"):
        row = snapshot["COMPACT_WAREHOUSE"]
        return {"code": str(_field(row, "WhsCode")), "name": str(_field(row, "WhsName", default=store_reference))}
    reference = str(store_reference or "").strip().lower()
    warehouses = snapshot.get("OWHS", [])
    exact = [row for row in warehouses if str(_field(row, "WhsCode", default="")).lower() == reference or str(_field(row, "WhsName", default="")).lower() == reference]
    candidates = exact or [row for row in warehouses if reference and reference in str(_field(row, "WhsName", default="")).lower()]
    if len(candidates) != 1:
        raise ValueError("门店与Core仓库无法唯一匹配，请先校准门店编码")
    return {
        "code": str(_field(candidates[0], "WhsCode")),
        "name": str(_field(candidates[0], "WhsName", default=store_reference)),
    }


def _catalog(snapshot):
    if "COMPACT_PRODUCTS" in snapshot:
        return {
            str(_field(row, "ItemCode")): {
                "product_code": str(_field(row, "ItemCode")),
                "product_name": str(_field(row, "ItemName", default=_field(row, "ItemCode"))),
                "brand": str(_field(row, "FirmName", default="品牌待校准")),
                "category": str(_field(row, "ItmsGrpNam", default="未分类")),
            }
            for row in snapshot["COMPACT_PRODUCTS"] if _field(row, "ItemCode")
        }
    categories = {str(_field(row, "ItmsGrpCod")): str(_field(row, "ItmsGrpNam", default="未分类")) for row in snapshot.get("OITB", [])}
    brands = {str(_field(row, "FirmCode")): str(_field(row, "FirmName", default="品牌待校准")) for row in snapshot.get("OMRC", [])}
    products = {}
    for row in snapshot.get("OITM", []):
        code = str(_field(row, "ItemCode", default="")).strip()
        if not code:
            continue
        brand = _field(row, "U_Brand", "U_BrandName", default=None) or brands.get(str(_field(row, "FirmCode")), "品牌待校准")
        products[code] = {
            "product_code": code,
            "product_name": str(_field(row, "ItemName", default=code)),
            "brand": str(brand),
            "category": categories.get(str(_field(row, "ItmsGrpCod")), "未分类"),
        }
    return products


def _sales(snapshot, warehouse_code, as_of):
    if "COMPACT_SALES" in snapshot:
        return {
            str(_field(row, "ItemCode")): {
                "30": _number(_field(row, "Sales30")), "60": _number(_field(row, "Sales60")),
                "90": _number(_field(row, "Sales90")), "180": _number(_field(row, "Sales180")),
                "last_sale": _day(_field(row, "LastSaleDate")),
            }
            for row in snapshot["COMPACT_SALES"] if _field(row, "ItemCode")
        }
    documents = {}
    for row in snapshot.get("OINV", []):
        if str(_field(row, "CANCELED", "Canceled", default="N")).upper() == "Y":
            continue
        documents[("invoice", str(_field(row, "DocEntry")))] = _day(_field(row, "DocDate"))
    for row in snapshot.get("ORIN", []):
        if str(_field(row, "CANCELED", "Canceled", default="N")).upper() == "Y":
            continue
        documents[("return", str(_field(row, "DocEntry")))] = _day(_field(row, "DocDate"))
    sales = defaultdict(lambda: {"30": 0.0, "60": 0.0, "90": 0.0, "180": 0.0, "last_sale": None})
    for kind, table, sign in (("invoice", "INV1", 1), ("return", "RIN1", -1)):
        for row in snapshot.get(table, []):
            if str(_field(row, "WhsCode", default="")) != warehouse_code:
                continue
            item = str(_field(row, "ItemCode", default="")).strip()
            doc_date = documents.get((kind, str(_field(row, "DocEntry"))))
            if not item or not doc_date or doc_date > as_of:
                continue
            quantity = _number(_field(row, "Quantity")) * sign
            age = (as_of - doc_date).days
            for period in (30, 60, 90, 180):
                if age < period:
                    sales[item][str(period)] += quantity
            if sign > 0 and (sales[item]["last_sale"] is None or doc_date > sales[item]["last_sale"]):
                sales[item]["last_sale"] = doc_date
    return sales


def _inventory(snapshot, warehouse_code):
    if "COMPACT_PRODUCTS" in snapshot:
        compact = {}
        for row in snapshot["COMPACT_PRODUCTS"]:
            code = str(_field(row, "ItemCode", default="")).strip()
            if not code:
                continue
            compact[code] = {
                "current_inventory": _number(_field(row, "OnHand")),
                "committed_inventory": _number(_field(row, "IsCommited", "IsCommitted")),
                "incoming_inventory": _number(_field(row, "OnOrder")),
                "average_cost": _number(_field(row, "AvgPrice")),
            }
            compact[code]["sellable_inventory"] = max(0, compact[code]["current_inventory"] - compact[code]["committed_inventory"])
        return compact
    result = {}
    for row in snapshot.get("OITW", []):
        if str(_field(row, "WhsCode", default="")) != warehouse_code:
            continue
        code = str(_field(row, "ItemCode", default="")).strip()
        result[code] = {
            "current_inventory": _number(_field(row, "OnHand")),
            "committed_inventory": _number(_field(row, "IsCommited", "IsCommitted")),
            "incoming_inventory": _number(_field(row, "OnOrder")),
            "average_cost": _number(_field(row, "AvgPrice")),
        }
        result[code]["sellable_inventory"] = max(0, result[code]["current_inventory"] - result[code]["committed_inventory"])
    return result


def replenishment_analysis(snapshot, store_reference, as_of=None):
    as_of = as_of or date.today()
    warehouse = resolve_warehouse(snapshot, store_reference)
    products = _catalog(snapshot)
    inventory = _inventory(snapshot, warehouse["code"])
    sales = _sales(snapshot, warehouse["code"], as_of)
    items = []
    for code in sorted(set(products) | set(inventory) | set(sales)):
        product = products.get(code, {"product_code": code, "product_name": code, "brand": "品牌待校准", "category": "未分类"})
        stock = inventory.get(code, {"current_inventory": 0, "sellable_inventory": 0, "incoming_inventory": 0, "average_cost": 0})
        velocity = sales.get(code, {"30": 0, "60": 0, "90": 0, "180": 0, "last_sale": None})
        recent = max(0, velocity["30"])
        previous = max(0, velocity["60"] - velocity["30"])
        trend = "上升" if recent > previous * 1.1 and recent > 0 else "下降" if previous > recent * 1.1 else "平稳"
        target = recent
        suggested = _round_quantity(target - stock["sellable_inventory"] - stock["incoming_inventory"])
        reason = "近30天销售{:.0f}件，可售库存{:.0f}件，在途{:.0f}件。".format(recent, stock["sellable_inventory"], stock["incoming_inventory"])
        if suggested:
            reason += "按30天需求口径，建议补货{}件，需采购员人工确认。".format(suggested)
        else:
            reason += "现有库存可覆盖目标需求，暂不建议补货。"
        items.append({**product, **stock, "sales_30": round(velocity["30"], 2), "sales_60": round(velocity["60"], 2), "sales_90": round(velocity["90"], 2), "sales_trend": trend, "suggested_quantity": suggested, "ai_reason": reason})
    return {"store": warehouse, "as_of": as_of.isoformat(), "items": items, "source_tables": list(CORE_TABLES), "human_confirmation_required": True}


def inventory_health_analysis(snapshot, store_reference, as_of=None):
    as_of = as_of or date.today()
    warehouse = resolve_warehouse(snapshot, store_reference)
    products = _catalog(snapshot)
    inventory = _inventory(snapshot, warehouse["code"])
    sales = _sales(snapshot, warehouse["code"], as_of)
    items = []
    levels = {"健康": 0, "黄色": 0, "橙色": 0, "红色": 0, "黑色": 0}
    for code, stock in inventory.items():
        if stock["current_inventory"] <= 0:
            continue
        product = products.get(code, {"product_code": code, "product_name": code, "brand": "品牌待校准", "category": "未分类"})
        velocity = sales.get(code, {"90": 0, "180": 0, "last_sale": None})
        last_sale = velocity.get("last_sale")
        days = (as_of - last_sale).days if last_sale else 9999
        level = "黑色" if days >= 360 else "红色" if days >= 270 else "橙色" if days >= 180 else "黄色" if days >= 90 else "健康"
        levels[level] += 1
        advice = "保持正常销售观察。"
        if level == "黄色":
            advice = "复核陈列与员工产品讲解，连续观察30天。"
        elif level == "橙色":
            advice = "优先调整陈列、搭配销售并安排活动推广。"
        elif level == "红色":
            advice = "评估跨店调拨和专项推广，提交负责人确认。"
        elif level == "黑色":
            advice = "列入重点积压清单，复核成本、品牌规则后再决定促销处理。"
        items.append({**product, "inventory_quantity": round(stock["current_inventory"], 2), "inventory_amount": round(stock["current_inventory"] * stock["average_cost"], 2), "last_sale_date": last_sale.isoformat() if last_sale else None, "inventory_days": days, "sales_90": round(velocity.get("90", 0), 2), "sales_180": round(velocity.get("180", 0), 2), "health_level": level, "ai_advice": advice})
    items.sort(key=lambda row: (row["inventory_days"], row["inventory_amount"]), reverse=True)
    return {"store": warehouse, "as_of": as_of.isoformat(), "items": items, "levels": levels, "inventory_amount": round(sum(item["inventory_amount"] for item in items), 2), "source_tables": list(CORE_TABLES), "human_confirmation_required": True}


def _save_workbook(headers, rows, output_path, title):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="176443")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(42, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(output_path)
    return str(output_path)


def export_replenishment(result, output_path):
    headers = ["门店", "商品编码", "商品名称", "品牌", "类别", "当前库存", "30天销量", "60天销量", "90天销量", "销售趋势", "建议补货数量", "AI原因"]
    rows = [[result["store"]["name"], item["product_code"], item["product_name"], item["brand"], item["category"], item["current_inventory"], item["sales_30"], item["sales_60"], item["sales_90"], item["sales_trend"], item["suggested_quantity"], item["ai_reason"]] for item in result["items"]]
    return _save_workbook(headers, rows, output_path, "智能补货建议")


def export_inventory_health(result, output_path):
    headers = ["门店", "品牌", "商品编码", "商品", "库存数量", "库存金额", "最后销售日期", "库存天数", "90天销量", "180天销量", "风险等级", "AI建议"]
    rows = [[result["store"]["name"], item["brand"], item["product_code"], item["product_name"], item["inventory_quantity"], item["inventory_amount"], item["last_sale_date"] or "无销售记录", item["inventory_days"] if item["inventory_days"] < 9999 else "无销售记录", item["sales_90"], item["sales_180"], item["health_level"], item["ai_advice"]] for item in result["items"]]
    return _save_workbook(headers, rows, output_path, "积压库存分析")


def export_root():
    return Path(os.environ.get("OPERATION_EXPORT_ROOT", "/data/operation_exports"))


def source_timestamp():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")
